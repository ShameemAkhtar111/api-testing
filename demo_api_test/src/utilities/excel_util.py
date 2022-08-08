import openpyxl


class ExcelUtil:
    __filePath = None
    __workBook = None
    __sheet = None
    __row = None
    __col = None

    def __init__(self, file_path):
        if file_path is not None:
            if ".xlsx" in file_path:
                self.__filePath = file_path
                self.__workBook = openpyxl.load_workbook(self.__filePath)
                self.__sheet = self.__workBook.active
                self.__row = self.__sheet.max_row
                self.__col = self.__sheet.max_column
            else:
                raise Exception("Please provide full file path i.e till file_name.xlsx")
        else:
            raise Exception(f"Excel file path must be provided")

    def get_row_and_column_count(self, sheet_name=None):
        if sheet_name is not None:
            self.__sheet = self.__workBook[sheet_name]
            self.__row = self.__sheet.max_row
            self.__col = self.__sheet.max_column
        return self.__sheet.max_row, self.__sheet.max_column

    def get_active_sheet_name(self):
        return self.__sheet.title

    def get_all_sheet_names(self):
        return self.__workBook.sheetnames

    def get_cell_data(self, row_num, col_num, sheet_name=None):
        if sheet_name is not None:
            self.__sheet = self.__workBook[sheet_name]
        return self.__sheet.cell(row=row_num, column=col_num).value

    def get_column_num(self, col_name, sheet_name=None):
        col_index = -1
        if sheet_name is not None:
            self.__row, self.__col = self.get_row_and_column_count(sheet_name)

        for col in self.__sheet.iter_cols(min_row=1, max_col=self.__col, max_row=1):
            for cell in col:
                if cell.value == col_name:
                    col_index = cell.col_idx
                    break
        return col_index

    def get_cell_data_based_on_col_name(self, col_name, row_num, sheet_name=None):
        col_index = self.get_column_num(col_name, sheet_name)
        if col_index == -1:
            return f"No value found for column: {col_name} and row num: {row_num}"
        cell_value = self.get_cell_data(row_num, col_index, sheet_name)
        return cell_value

    def set_cell_data(self, row_num, col_num, data_to_write, sheet_name=None):
        if sheet_name is not None:
            self.get_row_and_column_count()
        cellToWrite = self.__sheet.cell(row_num, col_num)
        cellToWrite.value = data_to_write
        self.__workBook.save(self.__filePath)

if __name__ == '__main__':
    obj = ExcelUtil(r"C:\Users\Shameem Akhtar\Downloads\Compressed\Olympic Athletes.xlsx")
    print(obj.get_row_and_column_count("OlympicAthletes"))
    print(obj.get_active_sheet_name())
    print(obj.get_all_sheet_names())
    print(obj.get_cell_data(1, 2))
    print(obj.get_column_num("Sport"))
    print(obj.get_cell_data_based_on_col_name("Sport", 6))
    print(obj.set_cell_data(8620, 6,"hey"))
